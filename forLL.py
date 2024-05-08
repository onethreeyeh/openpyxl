from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border,Side

llshw = load_workbook("llshw.xlsx",data_only=True)

zhsheet = llshw["工作表"]
ygssheet = llshw["医工所项目表"]
hzwsheet = llshw["杭州湾项目表"]

#团队负责人
TeamLeader = ["许高杰","曹彦伟","曹鸿涛","刘兆平","诸葛飞","房丰洲","原子尺度与微纳制造实验室（筹）","江南","王立平","汪爱英","宋振纶",
"曾志翔","张涛","茅东升","蒲吉斌","黄政仁","陈涛","刘小青","常可可","海洋实验室","朱锦","郑文革","姚强",
"方省众","陈鹏","刘富","张永刚","王震","张浩","张建","余海斌","高分子实验室","黄庆","张蕾","先进核能材料实验室","王军强","李国伟",
"李润伟","磁材实验室","满其奎","磁性材料应用技术研究中心","闫阿儒","稀土永磁材料联合创新中心","张驰","祝颖丹","蒋俊","肖江剑","赵夙","机器人实验室",
"张文武","韦超","激光极端制造研究中心","陈新民","特飞中心","宋伟杰","叶继春","向超宇","葛子义","光电实验室","陈亮","况永波","尹宏峰","张亚杰","姚霞银","官万兵","杨明辉",
"夏永高","陆之毅","谢银君","氢能实验室","赵一天","张建涛","吴爱国","李华","左国坤","王荣","郑建萍","郭明全",
"李辉","先进诊疗实验室","平台"]

#23年到位
data23 = [[5.61,98.81,0,0,0,344.97,0,127,0,47],
[191.28,445,0,20,35,165,0,0,12.74,224.9],
[42.32,109.32,0,90,44,0,120,0,0,60.82],
[28.17,67.5,72.45,0,62,5,0,0,0,0],
[89.39,0,0,0,0,6,0,0,0,0],
[0,0,0,0,0,0,0,0,0,0],
[0,0,0,0,0,0,0,0,0,0],
[60.47,247.01,0,16.95,0,408.5,45.23,1437.28,30,283.83],
[1036.13,503.32,0,840.1,35,440.5,90,1022.67,130.18,795.64],
[149.5,41.05,0,0,20,252,0,11.02,0,127.2],
[2.04,44,0,0,30,200,60,0,50,0],
[50,0,0,0,0,45,0,43,115.47,0],
[126.08,0,0,0,35,78,0,0,6,0],
[18,0,0,0,0,27,31.99,0,0,0],
[267,248.26,0,0,410,445,0,300.02,18,327.76],
[188,1334.1,0,0,0,0,0,345.89,0,955],
[233.71,161.24,0,0,90,192,60,0,15,49.85],
[233.08,152,0,0,10,150,0,260,0,38],
[39,364.6,0,54,0,5,0,0,0,322],
[0,0,0,75,0,131.98,0,0,0,0],
[68.73,603.88,0,0,103.4,299.8,0,0,114.05,725.94],
[24.16,0,0,0,256,364.5,0,33.02,0,6],
[0,0,0,0,0,0,0,0,0,0],
[0,24.19,0,0,0,0,25,0,29.5,0],
[0,0,0,0,33,133.3,0,0,6.25,0],
[43,0,0,0,168,95,256.22,0,0,110.9],
[0,0,0,0,316,100,0,382.407,0,63],
[12,0,0,855,70,36,0,3542.07,0,141.75],
[0,0,0,0,151.2,200,0,352.54,0,30],
[209.75,0,0,0,20,145,0,0,0,0],
[0,0,0,0,0,80,0,0,6,46.8],
[0,0,0,0,0,0,0,0,0,0],
[396.76,49,0,0,480,105,0,13.27,0,45.94],
[124.32,0,0,0,0,10,0,74.65,0,0],
[0,0,0,0,0,0,0,0,0,0],
[75.38,0,0,0,10,11,0,0,0,0],
[7.05,0,0,0,44,0,0,0,7,33.6],
[781.97,14.4,0,60,307.6,166,60.61,24,0,210.81],
[0,0,0,0,0,0,0,0,0,0],
[55.94,101.75,0,0,400,180,0,0,0,176.3],
[0,0,0,0,0,0,0,0,0,0],
[54.1,2180.39,0,0,216.35,557,0,1501,66.07,1261.69],
[0,0,0,0,0,0,0,0,0,0],
[639.78,324,18.37,1369.51,677.4,967.5,102.09,160.93,0,1054.4],
[16.549,15,0,0,0,341,0,96.97,14.25,124.76],
[84.12,0,0,0,150,120,60,135.93,0,0],
[0,0,0,0,10,381.76,0,294.69,0,64.65],
[0,0,0,0,0,5,0,0,0,0],
[0,0,0,0,0,0,0,0,0,0],
[44,0,0,0,0,336,0,57.06,0,124.44],
[0,51.7,0,0,0,5,60,70.92,0,0],
[0,0,0,0,0,0,0,0,0,0],
[0,0,0,20,300,165,0,100,0,0],
[0,0,0,0,0,0,0,0,0,0],
[0,115.85,0,40,0,0,0,18,0,0],
[63.22,48,0,0,115,125,0,0,0,72],
[0,414.33,0,0,35,3.2,0,0,0,153.48],
[153.1,0,0,0,0,140,0,50,0,30],
[0,0,0,0,0,0,0,0,0,0],
[58.85,0,0,60,79,125,0,360,8,3],
[36.05,0,0,0,15,97,45,0,0,0],
[0,0,0,0,0,45,0,0,0,10.747],
[0,85.1,0,0,0,30,0,0,0,168],
[110.05,34.36,0,0,0,60,12,90,90,0],
[99.75,701.6,0,0,92,0,0,49.15,0,394.17],
[2.95,0,0,0,0,0,0,0,0,0],
[15.75,39.85,100,0,103.68,204,75.42,98,10.85,8],
[56,0,0,51.78,0,159,0,0,0,0],
[0,0,0,0,0,60,0,0,0,0],
[0,0,0,0,0,0,0,0,0,0],
[70.22,0,0,0,33,87,0,15.18,5,0],
[12,0,0,0,15,22,0,0,0,0],
[150,139,0,24.6,10.9,470,0,0,0,8],
[14.75,0,0,0,0,72,0,134.04,0,30],
[0,96,0,0,177,132,0,0,8,111],
[2.04,0,0,0,0,8.5,0,0,0,0],
[0,162.8,0,0,0,21.5,0,0,0,56.27],
[0,0,0,0,0,0,0,0,0,0],
[0,0,0,0,0,0,0,0,0,0],
[0,0,0,0,700,120,0,0,0,27.98],
[12,0,2000,0,0,463.83,0,0,0,0]]


#创建空数组
data_rows = 243
data_cols = 10
data = []
for i in range(data_rows):
	if(i%3 == 2):
		data.append(data23[int(i/3)])
	else:
		row = []
		for j in range(data_cols):
			row.append(0)
		data.append(row)

for row in zhsheet:
	if(row[74].value == "科技发展部" or row[74].value == "科技发展部（重任处）"):
		#高技术
		if(row[73].value == "高技术"):
			#到位经费
			if (row[64].value == "YES"):
				TeamIndex = TeamLeader.index(row[67].value)
				if(row[62].value != None):
					data[3*TeamIndex + 1][7] += row[62].value
				if(row[63].value != None):
					data[3*TeamIndex + 1][9] += row[63].value
			#合同经费
			if(row[7].value == 24):
				TeamIndex = TeamLeader.index(row[67].value)
				data[3*TeamIndex][7] += row[8].value
		#国际合作
		elif(row[73].value == "国际（地区）合作与交流项目" or \
			row[73].value == "国际工程科技高端论坛" or \
			row[73].value == "国际合作" or row[73].value == "国际会议" or \
			row[73].value == "国际交流" or row[73].value == "国际人才" or \
			row[73].value == "国际载体" or row[73].value == "海外及港澳学者合作研究基金" or \
			row[73].value == "外国青年学者项目" or row[73].value == "外国学者研究基金(资深项目)"):
			#到位经费
			if (row[64].value == "YES"):
				TeamIndex = TeamLeader.index(row[67].value)
				if(row[62].value != None):
					data[3*TeamIndex + 1][6] += row[62].value
				if(row[63].value != None):
					data[3*TeamIndex + 1][9] += row[63].value
			#合同经费
			if(row[7].value == 24):
				TeamIndex = TeamLeader.index(row[67].value)
				data[3*TeamIndex][6] += row[8].value
		#其他
		else:
			#国家
			if(row[71].value == "国家"):
				#国家基金
				if(row[72].value == "国家自然科学基金"):
					#到位经费
					if(row[64].value == "YES"):
						TeamIndex = TeamLeader.index(row[67].value)
						if(row[62].value != None):
							data[3*TeamIndex + 1][0] += row[62].value
						if(row[63].value != None):
							data[3*TeamIndex + 1][9] += row[63].value
					if(row[7].value == 24):
						TeamIndex = TeamLeader.index(row[67].value)
						data[3*TeamIndex][0] += row[8].value
				#科技部
				elif(row[72].value == "国家重点研发计划" or\
					row[72].value == "国家重点研发计划（华北电力大学）"or\
					row[72].value == "国家科技部" or\
					row[72].value == "科技部" or\
					row[72].value == "科技部杰出青年科学家项目"):
					#到位经费
					if(row[64].value == "YES"):
						TeamIndex = TeamLeader.index(row[67].value)
						if(row[62].value != None):
							data[3*TeamIndex + 1][1] += row[62].value
						if(row[63].value != None):
							data[3*TeamIndex + 1][9] += row[63].value
					if(row[7].value == 24 and row[8].value != None):
						TeamIndex = TeamLeader.index(row[67].value)
						data[3*TeamIndex][1] += row[8].value
				#国家其他
				else:
					#到位经费
					if(row[64].value == "YES"):
						TeamIndex = TeamLeader.index(row[67].value)
						if(row[62].value != None):
							data[3*TeamIndex + 1][2] += row[62].value
						if(row[63].value != None):
							data[3*TeamIndex + 1][9] += row[63].value
					if(row[7].value == 24 and row[8].value != None):
						TeamIndex = TeamLeader.index(row[67].value)
						data[3*TeamIndex][2] += row[8].value
			#中科院
			elif(row[71].value == "中科院"):
				#到位经费
				if (row[64].value == "YES"):
					TeamIndex = TeamLeader.index(row[67].value)
					if(row[62].value != None):
						data[3*TeamIndex + 1][3] += row[62].value
					if(row[63].value != None):
						data[3*TeamIndex + 1][9] += row[63].value
				#合同经费
				if(row[7].value == 24):
					TeamIndex = TeamLeader.index(row[67].value)
					data[3*TeamIndex][3] += row[8].value
			#浙江省
			elif(row[71].value == "浙江省"):
				#到位经费
				if (row[64].value == "YES"):
					TeamIndex = TeamLeader.index(row[67].value)
					if(row[62].value != None):
						data[3*TeamIndex + 1][4] += row[62].value
					if(row[63].value != None):
						data[3*TeamIndex + 1][9] += row[63].value
				#合同经费
				if(row[7].value == 24):
					TeamIndex = TeamLeader.index(row[67].value)
					data[3*TeamIndex][4] += row[8].value
			#宁波市
			elif(row[71].value == "宁波市"):
				#到位经费
				if (row[64].value == "YES"):
					TeamIndex = TeamLeader.index(row[67].value)
					if(row[62].value != None):
						data[3*TeamIndex + 1][5] += row[62].value
					if(row[63].value != None):
						data[3*TeamIndex + 1][9] += row[63].value
				#合同经费
				if(row[7].value == 24):
					TeamIndex = TeamLeader.index(row[67].value)
					data[3*TeamIndex][5] += row[8].value
			else:
				#到位经费
				if (row[64].value == "YES"):
					TeamIndex = TeamLeader.index(row[67].value)
					if(row[62].value != None):
						data[3*TeamIndex + 1][8] += row[62].value
					if(row[63].value != None):
						data[3*TeamIndex + 1][9] += row[63].value
				#合同经费
				if(row[7].value == 24):
					TeamIndex = TeamLeader.index(row[67].value)
					data[3*TeamIndex][8] += row[8].value

for row in ygssheet:
	if(row[44].value == "科技发展部" or row[44].value == "科技发展部（重任处）"):
		#高技术
		if(row[43].value == "高技术"):
			#到位经费
			if(row[35].value == "YES"):
				TeamIndex = TeamLeader.index(row[37].value)
				if(row[33].value != None):
					data[3*TeamIndex + 1][7] += row[33].value
				if(row[34].value != None):
					data[3*TeamIndex + 1][9] += row[34].value
			#合同经费
			if(row[6].value == 24):
				TeamIndex = TeamLeader.index(row[37].value)
				data[3*TeamIndex][7] += row[7].value
		#其他
		else:
			#浙江省
			if(row[41].value == "浙江省"):
				#到位经费
				if(row[35].value == "YES"):
					TeamIndex = TeamLeader.index(row[37].value)
					if(row[33].value != None):
						data[3*TeamIndex + 1][4] += row[33].value
					if(row[34].value != None):
						data[3*TeamIndex + 1][9] += row[34].value
				#合同经费
				if(row[6].value == 24):
					TeamIndex = TeamLeader.index(row[37].value)
					data[3*TeamIndex][4] += row[7].value
			#宁波市
			elif(row[41].value == "宁波市"):
				#到位经费
				if(row[35].value == "YES"):
					TeamIndex = TeamLeader.index(row[37].value)
					if(row[33].value != None):
						data[3*TeamIndex + 1][5] += row[33].value
					if(row[34].value != None):
						data[3*TeamIndex + 1][9] += row[34].value
				#合同经费
				if(row[6].value == 24):
					TeamIndex = TeamLeader.index(row[37].value)
					data[3*TeamIndex][5] += row[7].value
			#其他
			else:
				#到位经费
				if(row[35].value == "YES"):
					TeamIndex = TeamLeader.index(row[37].value)
					if(row[33].value != None):
						data[3*TeamIndex + 1][8] += row[33].value
					if(row[34].value != None):
						data[3*TeamIndex + 1][9] += row[34].value
				#合同经费
				if(row[6].value == 24):
					TeamIndex = TeamLeader.index(row[37].value)
					data[3*TeamIndex][8] += row[7].value


for row in hzwsheet:
	if(row[44].value == "科技发展部" or row[44].value == "科技发展部（重任处）"):
		if(row[41].value == "浙江省"):
			#到位经费
			if(row[35].value == "YES"):
				TeamIndex = TeamLeader.index(row[37].value)
				if(row[33].value != None):
					data[3*TeamIndex + 1][4] += row[33].value
				if(row[34].value != None):
					data[3*TeamIndex + 1][9] += row[34].value
			#合同经费
			if(row[6].value == 24):
				TeamIndex = TeamLeader.index(row[37].value)
				data[3*TeamIndex][4] += row[7].value
		#宁波市
		elif(row[41].value == "宁波市"):
			#到位经费
			if(row[35].value == "YES"):
				TeamIndex = TeamLeader.index(row[37].value)
				if(row[33].value != None):
					data[3*TeamIndex + 1][5] += row[33].value
				if(row[34].value != None):
					data[3*TeamIndex + 1][9] += row[34].value
			#合同经费
			if(row[6].value == 24):
				TeamIndex = TeamLeader.index(row[37].value)
				data[3*TeamIndex][5] += row[7].value
		#其他
		else:
			#到位经费
			if(row[35].value == "YES"):
				TeamIndex = TeamLeader.index(row[37].value)
				if(row[33].value != None):
					data[3*TeamIndex + 1][8] += row[33].value
				if(row[34].value != None):
					data[3*TeamIndex + 1][9] += row[34].value
			#合同经费
			if(row[6].value == 24):
				TeamIndex = TeamLeader.index(row[37].value)
				data[3*TeamIndex][8] += row[7].value	


#画表
answer = Workbook()
sheet = answer["Sheet"]
sheet["A1"].value = "团队名称"
sheet["B1"].value = "经费类别"
sheet["C1"].value = "国家基金"
sheet["D1"].value = "科技部"
sheet["E1"].value = "国家其他"
sheet["F1"].value = "中科院"
sheet["G1"].value = "浙江省"
sheet["H1"].value = "宁波市"
sheet["I1"].value = "国际合作"
sheet["J1"].value = "高技术"
sheet["K1"].value = "其他"
sheet["L1"].value = "转出经费"

for i in range(81):
	sheet.cell(3*i+2,1,TeamLeader[i])

for i in range(2,245):
	if(i%3 == 2):
		sheet.cell(i,2,"2024年合同")
	elif(i%3 == 0):
		sheet.cell(i,2,"2024年到位")
	else:
		sheet.cell(i,2,"2023年到位")

for i in range(243):
	for j in range(10):
		if(data[i][j] == 0):
			sheet.cell(i+2,j+3,None)
		else:
			sheet.cell(i+2,j+3,data[i][j])


#设置边框
border = Border(left=Side(border_style='thin'), 
                right=Side(border_style='thin'), 
                top=Side(border_style='thin'), 
                bottom=Side(border_style='thin'))
cell_range = "A1:P274"
for row in sheet[cell_range]:
	for cell in row:
		cell.border = border

answer.save("answer.xlsx")