from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border,Side

llshw = load_workbook("llshw.xlsx",data_only=True)

zhsheet = llshw["工作表"]
ygssheet = llshw["先进诊疗材料与技术实验室项目表"]
hzwsheet = llshw["杭州湾项目表"]

#创建团队负责人和到位经费数组
dw24 = load_workbook("dw24.xlsx",data_only = True)
dwsheet = dw24["到位经费"]
maxrow = dwsheet.max_row
TeamLeader = []
data24 = []
for row in dwsheet:
	if (row[0].value != None):
		TeamLeader.append(row[0].value)
		currentRow = []
		for i in range(1,11):
			if (row[i].value == None):
				currentRow.append(0)
			else:
				currentRow.append(row[i].value)
		data24.append(currentRow)

#本部园区各列
laiyuanfangxiang = 76
ketixingzhi = 75
jingfeizengjian = 67
fuzeren = 69
daoweijingfei = 65
zhuanchujingfei = 66
hetongjingfei = 8
ketilaiyuan  = 73
jingfeilaiyuan = 74

#医工所杭州湾各列
ygslaiyuanfangxiang = 47
ygsketixingzhi = 46
ygsfuzeren = 40
ygsjingfeizengjian = 38
ygsdaoweijingfei = 36
ygszhuanchujingfei = 37
ygsketilaiyuan = 44



#创建空数组
data_rows = (len(data24))*3
data_cols = 10
data = []
for i in range(data_rows):
	if(i%3 == 2):
		data.append(data24[int(i/3)])
	else:
		row = []
		for j in range(data_cols):
			row.append(0)
		data.append(row)

for row in zhsheet:
	if(row[laiyuanfangxiang].value == "科技发展部" or row[laiyuanfangxiang].value == "科技发展部（重任处）" or row[laiyuanfangxiang].value == "科技发展部（科技处）"):
		#高技术
		if(row[ketixingzhi].value == "高技术" or row[ketixingzhi].value == "高技术（重大）"):
			#到位经费
			if (row[jingfeizengjian].value == "YES"):
				TeamIndex = TeamLeader.index(row[fuzeren].value)
				if(row[daoweijingfei].value != None):
					data[3*TeamIndex + 1][7] += row[daoweijingfei].value
				if(row[zhuanchujingfei].value != None):
					data[3*TeamIndex + 1][9] += row[zhuanchujingfei].value
			#合同经费
			if(row[7].value == 25):
				TeamIndex = TeamLeader.index(row[fuzeren].value)
				data[3*TeamIndex][7] += row[hetongjingfei].value
		#国际合作
		elif(row[ketixingzhi].value == "国际（地区）合作与交流项目" or \
			row[ketixingzhi].value == "国际工程科技高端论坛" or \
			row[ketixingzhi].value == "国际合作" or row[ketixingzhi].value == "国际会议" or \
			row[ketixingzhi].value == "国际交流" or row[ketixingzhi].value == "国际人才" or \
			row[ketixingzhi].value == "国际载体" or row[ketixingzhi].value == "海外及港澳学者合作研究基金" or \
			row[ketixingzhi].value == "外国青年学者项目" or row[ketixingzhi].value == "外国学者研究基金(资深项目)"):
			#到位经费
			if (row[jingfeizengjian].value == "YES"):
				TeamIndex = TeamLeader.index(row[fuzeren].value)
				if(row[daoweijingfei].value != None):
					data[3*TeamIndex + 1][6] += row[daoweijingfei].value
				if(row[zhuanchujingfei].value != None):
					data[3*TeamIndex + 1][9] += row[zhuanchujingfei].value
			#合同经费
			if(row[7].value == 25):
				TeamIndex = TeamLeader.index(row[fuzeren].value)
				data[3*TeamIndex][6] += row[hetongjingfei].value
		#其他
		else:
			#国家
			if(row[ketilaiyuan].value == "国家"):
				#国家基金
				if(row[jingfeilaiyuan].value == "国家自然科学基金"):
					#到位经费
					if(row[jingfeizengjian].value == "YES"):
						TeamIndex = TeamLeader.index(row[fuzeren].value)
						if(row[daoweijingfei].value != None):
							data[3*TeamIndex + 1][0] += row[daoweijingfei].value
						if(row[zhuanchujingfei].value != None):
							data[3*TeamIndex + 1][9] += row[zhuanchujingfei].value
					if(row[7].value == 25 and row[hetongjingfei].value != None):
						TeamIndex = TeamLeader.index(row[fuzeren].value)
						data[3*TeamIndex][0] += row[hetongjingfei].value
				#科技部
				elif(row[jingfeilaiyuan].value == "国家重点研发计划" or\
					row[jingfeilaiyuan].value == "国家重点研发计划（华北电力大学）"or\
					row[jingfeilaiyuan].value == "国家科技部" or\
					row[jingfeilaiyuan].value == "科技部" or\
					row[jingfeilaiyuan].value == "科技部杰出青年科学家项目"):
					#到位经费
					if(row[jingfeizengjian].value == "YES"):
						TeamIndex = TeamLeader.index(row[fuzeren].value)
						if(row[daoweijingfei].value != None):
							data[3*TeamIndex + 1][1] += row[daoweijingfei].value
						if(row[zhuanchujingfei].value != None):
							data[3*TeamIndex + 1][9] += row[zhuanchujingfei].value
					if(row[7].value == 25 and row[hetongjingfei].value != None):
						TeamIndex = TeamLeader.index(row[fuzeren].value)
						data[3*TeamIndex][1] += row[hetongjingfei].value
				#国家其他
				else:
					#到位经费
					if(row[jingfeizengjian].value == "YES"):
						TeamIndex = TeamLeader.index(row[fuzeren].value)
						if(row[daoweijingfei].value != None):
							data[3*TeamIndex + 1][2] += row[daoweijingfei].value
						if(row[zhuanchujingfei].value != None):
							data[3*TeamIndex + 1][9] += row[zhuanchujingfei].value
					if(row[7].value == 25 and row[hetongjingfei].value != None):
						TeamIndex = TeamLeader.index(row[fuzeren].value)
						data[3*TeamIndex][2] += row[hetongjingfei].value
			#中科院
			elif(row[ketilaiyuan].value == "中科院"):
				#到位经费
				if (row[jingfeizengjian].value == "YES"):
					TeamIndex = TeamLeader.index(row[fuzeren].value)
					if(row[daoweijingfei].value != None):
						data[3*TeamIndex + 1][3] += row[daoweijingfei].value
					if(row[zhuanchujingfei].value != None):
						data[3*TeamIndex + 1][9] += row[zhuanchujingfei].value
				#合同经费
				if(row[7].value == 25):
					TeamIndex = TeamLeader.index(row[fuzeren].value)
					if(row[hetongjingfei].value != None):
						data[3*TeamIndex][3] += row[hetongjingfei].value
			#浙江省
			elif(row[ketilaiyuan].value == "浙江省"):
				#到位经费
				if (row[jingfeizengjian].value == "YES"):
					TeamIndex = TeamLeader.index(row[fuzeren].value)
					if(row[daoweijingfei].value != None):
						data[3*TeamIndex + 1][4] += row[daoweijingfei].value
					if(row[zhuanchujingfei].value != None):
						data[3*TeamIndex + 1][9] += row[zhuanchujingfei].value
				#合同经费
				if(row[7].value == 25):
					TeamIndex = TeamLeader.index(row[fuzeren].value)
					if(row[hetongjingfei].value != None):
						data[3*TeamIndex][4] += row[hetongjingfei].value
			#宁波市
			elif(row[ketilaiyuan].value == "宁波市"):
				#到位经费
				if (row[jingfeizengjian].value == "YES"):
					TeamIndex = TeamLeader.index(row[fuzeren].value)
					if(row[daoweijingfei].value != None):
						data[3*TeamIndex + 1][5] += row[daoweijingfei].value
					if(row[zhuanchujingfei].value != None):
						data[3*TeamIndex + 1][9] += row[zhuanchujingfei].value
				#合同经费
				if(row[7].value == 25):
					TeamIndex = TeamLeader.index(row[fuzeren].value)
					data[3*TeamIndex][5] += row[hetongjingfei].value
			elif(row[ketilaiyuan].value != "所内"):
				#到位经费
				if (row[jingfeizengjian].value == "YES"):
					TeamIndex = TeamLeader.index(row[fuzeren].value)
					if(row[daoweijingfei].value != None):
						data[3*TeamIndex + 1][8] += row[daoweijingfei].value
					if(row[zhuanchujingfei].value != None):
						data[3*TeamIndex + 1][9] += row[zhuanchujingfei].value
				#合同经费
				if(row[7].value == 25):
					TeamIndex = TeamLeader.index(row[fuzeren].value)
					data[3*TeamIndex][8] += row[hetongjingfei].value

for row in ygssheet:
	if(row[ygslaiyuanfangxiang].value == "科技发展部" or row[ygslaiyuanfangxiang].value == "科技发展部（重任处）" or row[ygslaiyuanfangxiang].value == "科技发展部（科技处）"):
		#高技术
		if(row[ygsketixingzhi].value == "高技术"):
			#到位经费
			if(row[ygsjingfeizengjian].value == "YES"):
				TeamIndex = TeamLeader.index(row[ygsfuzeren].value)
				if(row[ygsdaoweijingfei].value != None):
					data[3*TeamIndex + 1][7] += row[ygsdaoweijingfei].value
				if(row[ygszhuanchujingfei].value != None):
					data[3*TeamIndex + 1][9] += row[ygszhuanchujingfei].value
			#合同经费
			if(row[6].value == 25):
				TeamIndex = TeamLeader.index(row[ygsfuzeren].value)
				data[3*TeamIndex][7] += row[7].value
		#其他
		else:
			#浙江省
			if(row[ygsketilaiyuan].value == "浙江省"):
				#到位经费
				if(row[ygsjingfeizengjian].value == "YES"):
					TeamIndex = TeamLeader.index(row[ygsfuzeren].value)
					if(row[ygsdaoweijingfei].value != None):
						data[3*TeamIndex + 1][4] += row[ygsdaoweijingfei].value
					if(row[ygszhuanchujingfei].value != None):
						data[3*TeamIndex + 1][9] += row[ygszhuanchujingfei].value
				#合同经费
				if(row[6].value == 25):
					TeamIndex = TeamLeader.index(row[ygsfuzeren].value)
					data[3*TeamIndex][4] += row[7].value
			#宁波市
			elif(row[ygsketilaiyuan].value == "宁波市"):
				#到位经费
				if(row[ygsjingfeizengjian].value == "YES"):
					TeamIndex = TeamLeader.index(row[ygsfuzeren].value)
					if(row[ygsdaoweijingfei].value != None):
						data[3*TeamIndex + 1][5] += row[ygsdaoweijingfei].value
					if(row[ygszhuanchujingfei].value != None):
						data[3*TeamIndex + 1][9] += row[ygszhuanchujingfei].value
				#合同经费
				if(row[6].value == 25):
					TeamIndex = TeamLeader.index(row[ygsfuzeren].value)
					data[3*TeamIndex][5] += row[7].value
			#其他
			elif(row[ygsketilaiyuan].value != "所内"):
				#到位经费
				if(row[ygsjingfeizengjian].value == "YES"):
					TeamIndex = TeamLeader.index(row[ygsfuzeren].value)
					if(row[ygsdaoweijingfei].value != None):
						data[3*TeamIndex + 1][8] += row[ygsdaoweijingfei].value
					if(row[ygszhuanchujingfei].value != None):
						data[3*TeamIndex + 1][9] += row[ygszhuanchujingfei].value
				#合同经费
				if(row[6].value == 25):
					TeamIndex = TeamLeader.index(row[ygsfuzeren].value)
					data[3*TeamIndex][8] += row[7].value


for row in hzwsheet:
	if(row[ygslaiyuanfangxiang].value == "科技发展部" or row[ygslaiyuanfangxiang].value == "科技发展部（重任处）" or row[ygslaiyuanfangxiang].value == "科技发展部（科技处）"):
		if(row[ygsketilaiyuan].value == "浙江省"):
			#到位经费
			if(row[ygsjingfeizengjian].value == "YES"):
				TeamIndex = TeamLeader.index(row[ygsfuzeren].value)
				if(row[ygsdaoweijingfei].value != None):
					data[3*TeamIndex + 1][4] += row[ygsdaoweijingfei].value
				if(row[ygszhuanchujingfei].value != None):
					data[3*TeamIndex + 1][9] += row[ygszhuanchujingfei].value
			#合同经费
			if(row[6].value == 25):
				TeamIndex = TeamLeader.index(row[ygsfuzeren].value)
				data[3*TeamIndex][4] += row[7].value
		#宁波市
		elif(row[ygsketilaiyuan].value == "宁波市"):
			#到位经费
			if(row[ygsjingfeizengjian].value == "YES"):
				TeamIndex = TeamLeader.index(row[ygsfuzeren].value)
				if(row[ygsdaoweijingfei].value != None):
					data[3*TeamIndex + 1][5] += row[ygsdaoweijingfei].value
				if(row[ygszhuanchujingfei].value != None):
					data[3*TeamIndex + 1][9] += row[ygszhuanchujingfei].value
			#合同经费
			if(row[6].value == 25):
				TeamIndex = TeamLeader.index(row[ygsfuzeren].value)
				data[3*TeamIndex][5] += row[7].value
		#其他
		elif(row[ygsketilaiyuan].value != "所内"):
			#到位经费
			if(row[ygsjingfeizengjian].value == "YES"):
				TeamIndex = TeamLeader.index(row[ygsfuzeren].value)
				if(row[ygsdaoweijingfei].value != None):
					data[3*TeamIndex + 1][8] += row[ygsdaoweijingfei].value
				if(row[ygszhuanchujingfei].value != None):
					data[3*TeamIndex + 1][9] += row[ygszhuanchujingfei].value
			#合同经费
			if(row[6].value == 25):
				TeamIndex = TeamLeader.index(row[ygsfuzeren].value)
				data[3*TeamIndex][8] += row[7].value	


#画表
answer = Workbook()
sheet = answer["Sheet"]
sheet["A1"].value = "团队名称"
sheet["B1"].value = "经费类别"
sheet["C1"].value = "国家基金"
sheet["D1"].value = "科技部"
sheet["E1"].value = "国家其他"
sheet["F1"].value = "中科院"
sheet["G1"].value = "浙江省"
sheet["H1"].value = "宁波市"
sheet["I1"].value = "国际合作"
sheet["J1"].value = "高技术"
sheet["K1"].value = "其他"
sheet["L1"].value = "转出经费"

for i in range(len(TeamLeader)):
	sheet.cell(3*i+2,1,TeamLeader[i])

for i in range(2,data_rows+2):
	if(i%3 == 2):
		sheet.cell(i,2,"2025年合同")
	elif(i%3 == 0):
		sheet.cell(i,2,"2025年到位")
	else:
		sheet.cell(i,2,"2024年到位")

for i in range(data_rows):
	for j in range(data_cols):
		if(data[i][j] == 0):
			sheet.cell(i+2,j+3,None)
		else:
			sheet.cell(i+2,j+3,data[i][j])


#设置边框
border = Border(left=Side(border_style='thin'), 
                right=Side(border_style='thin'), 
                top=Side(border_style='thin'), 
                bottom=Side(border_style='thin'))
cell_range = "A1:L262"
for row in sheet[cell_range]:
	for cell in row:
		cell.border = border

answer.save("answer.xlsx")